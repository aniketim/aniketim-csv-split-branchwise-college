import streamlit as st
import pandas as pd
from io import BytesIO
import os

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Page setup
st.set_page_config(page_title="CSV to Branch-wise Excel", layout="wide")
st.title("📑 CSV to Branch-wise Excel Generator")

# Upload section
st.markdown("### 📤 Upload your CSV files")
uploaded_files = st.file_uploader(
    label="Upload one or more CSV files (Hold Ctrl or Shift to select multiple)",
    type=["csv"],
    accept_multiple_files=True
)

if uploaded_files:
    for uploaded_file in uploaded_files:
        uploaded_filename = os.path.splitext(uploaded_file.name)[0]

        try:
            df = pd.read_csv(uploaded_file)

            if "Select Branch" not in df.columns:
                st.error(f"❌ Column 'Select Branch' not found in {uploaded_file.name}. Skipping.")
                continue

            # Clean and identify branches
            df["Select Branch"] = df["Select Branch"].astype(str).str.strip()
            branches = df["Select Branch"].unique()

            # Create branch count table
            branch_counts = df["Select Branch"].value_counts().reset_index()
            branch_counts.columns = ["Branch", "Count"]
            total = branch_counts["Count"].sum()
            total_row = pd.DataFrame([["Total", total]], columns=["Branch", "Count"])
            branch_counts = pd.concat([branch_counts, total_row], ignore_index=True)

            # Overview Info
            test_title = uploaded_filename
            subject_name = "<Enter subject here>"

            # Generate initial Excel
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                pd.DataFrame().to_excel(writer, sheet_name="Overview")  # Empty sheet first
                df.to_excel(writer, sheet_name="All Data", index=False)

                for branch in branches:
                    branch_df = df[df["Select Branch"] == branch]
                    sheet_name = branch[:31]  # Excel limit
                    branch_df.to_excel(writer, sheet_name=sheet_name, index=False)

            output.seek(0)

            # Add Overview Sheet Formatting
            workbook = load_workbook(output)
            overview_ws = workbook["Overview"]

            # Overview Header
            overview_ws["A1"] = "Overview"
            overview_ws["A1"].font = Font(bold=True, size=16)

            # Space below "Overview"
            overview_ws["A3"] = "Test Title:"
            overview_ws["B3"] = test_title

            overview_ws["A5"] = "Subject Name:"
            overview_ws["B5"] = subject_name

            # Heading for branch-wise count
            overview_ws["A7"] = "Branch-wise Count"
            overview_ws["A7"].font = Font(bold=True)

            # Table headers
            overview_ws.cell(row=8, column=1, value="Branch").font = Font(bold=True)
            overview_ws.cell(row=8, column=2, value="Count").font = Font(bold=True)

            # Branch counts
            for idx, row in branch_counts.iterrows():
                r = 9 + idx
                overview_ws.cell(row=r, column=1, value=row["Branch"])
                count_cell = overview_ws.cell(row=r, column=2, value=row["Count"])
                if row["Branch"] == "Total":
                    overview_ws.cell(row=r, column=1).font = Font(bold=True)
                    count_cell.font = Font(bold=True)

            # Adjust column widths
            for col in range(1, 3):
                col_letter = get_column_letter(col)
                overview_ws.column_dimensions[col_letter].width = 20

            # Final output
            final_output = BytesIO()
            workbook.save(final_output)
            final_output.seek(0)

            output_filename = f"BranchWise_{uploaded_filename}.xlsx"

            st.success(f"✅ Processed: {uploaded_file.name}")
            st.download_button(
                label=f"📥 Download: {output_filename}",
                data=final_output,
                file_name=output_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        except Exception as e:
            st.error(f"⚠️ Error processing {uploaded_file.name}: {e}")
